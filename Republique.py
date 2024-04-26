from logging import Handler
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import Workbook,load_workbook
import os
import pathlib
import pandas as pd
import numpy as np
import xlrd
import datetime 
from selenium.webdriver.common.keys import Keys
import exception
import logging

# file = pathlib.Path("Republic_Francaise.xlsx")
# if file.exists("D:\Jannat USb\Jannat\Republique Francaise.xlsx"):
path="./Republique_Francaise.xlsx"
if os.path.isfile(path):
    try:
        flag=0
        driver=webdriver.Chrome(ChromeDriverManager().install())
        driver.maximize_window()

        #9 times comment code needed
        driver.get("https://www.1jeune1solution.gouv.fr/emplois?contrats=E2")
        date_time = datetime.datetime.now()
        date_time=date_time.strftime("%d/%m/%y")
        c=2


        
        with open("DetailPagesCount.txt") as DPC:
            # firstline = f.readlines()[0].rstrip()
            # firstline=firstline.split(':')
            # currentDate=firstline[1]
            PagesCountRead = DPC.readlines()
        minRows = PagesCountRead[0].rstrip()
        minRows=minRows.split(':')
        CurrentRow=minRows[1]
        maxRows = PagesCountRead[1].rstrip()
        maxRows=maxRows.split(':')
        TotalRow=maxRows[1]
        DPC.close()

        if(int(CurrentRow) < int(TotalRow)):
                    
           

            for i in range(int(CurrentRow),int(TotalRow) +1):
                cellsData=sheetObj.cell(row=i,column=8)

                link=cellsData.value
                driver.get(link)
                #driver.get('https://www.1jeune1solution.gouv.fr/emplois/116TRRW')
                Description=""
                CompanyName=""
                knowledges=""
                Experience=""
                Formation=""

                Description=driver.find_elements_by_xpath("//p[contains(@class,'job-offer-view__description ng-star-inserted')]")
                # Description=driver.find_elements_by_xpath("//p[@class='job-offer-view__description ng-star-inserted'] | //p[@class='app-inner-html--show-list-icons job-offer-view__description ng-star-inserted']")
                if len(Description)>0:
                    Descriptions=Description[0].text
                else:
                    Descriptions=''
                CompanyName=driver.find_elements_by_xpath("//div[@id='job-offer']/div[@class='ng-star-inserted']/a | //div[@id='job-offer']//a")
                if len(CompanyName)>0:
                    CompanyName=CompanyName[0].text.replace("\n","").replace("Postuler sur ","").replace("Apply on ","")
                else:
                    CompanyName=''
                infoDivs=driver.find_elements_by_xpath("//div[@class='job-offer-view__details ng-star-inserted']")
                if len(infoDivs)>0:
                    for x in infoDivs:
                        if(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Expérience') | (x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Experience'):
                            Experience=x.find_element_by_xpath('./p').text
                        elif(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Savoirs et savoir-faire') | (x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Knowledge and know-how'):
                            knowledge=x.find_elements_by_xpath('./p')
                            if len(knowledge)>0:
                                knowledges=knowledge
                            knowledge=x.find_elements_by_xpath('./ul/li/p')
                            if len(knowledge)>0:
                                for j in knowledge: 
                                    if len(knowledges)==0:
                                        knowledges=j.text
                                    else:
                                        knowledges=knowledges +','+j.text
                            # knowledge=x.find_elements_by_xpath('./p | ./ul/li/p')
                            # if len(knowledge)>0:
                            #     for j in knowledge:
                            #         knowledges=knowledges +','+j.text
                            # else:
                            #     print("")

                        elif(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Formation') | (x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Training'):
                            Formation=x.find_element_by_xpath('./p | ./ul/li/p').text
                else:
                    Experience=""
                    knowledges=""
                    Formation=""

                sheetObj.cell(row=i,column=4).value=CompanyName
                sheetObj.cell(row=i,column=10).value=Descriptions
                sheetObj.cell(row=i,column=11).value=Experience
                sheetObj.cell(row=i,column=12).value=knowledges
                sheetObj.cell(row=i,column=13).value=Formation


                wb_ReadObj.save("Republique_Francaise.xlsx")
            minRowUpdate=i
            wb_ReadObj.close()
            fileforDate = open("DateFile.txt","w+")
            fileforDate.writelines("Date:"+date_time)
            fileforDate.close()
            DetailPagesCount = open("DetailPagesCount.txt","w")
            DetailPagesCount.write("Total Detail Pages Complete:"+str(minRowUpdate)+"\n")
            DetailPagesCount.write("Total Rows in Excel:"+str(TotalRow)+"\n")
            DetailPagesCount.close()
            driver.close()
        else:
            with open("DateFile.txt") as f:
                firstline = f.readlines()[0].rstrip()
                firstline=firstline.split(':')
                currentDate=firstline[1]
            f.close()
            # k=df.loc[df['Domain Name'] == "Achat / Comptabilité / Gestion"]
            # f=k['Production Date']
            # date_object = datetime.strptime('08/07/21', '%d/%m/%y')
            # print(date_object)
            # print(max(k['Production Date'], datetime.strptime('%dd/%mm/%YY')))
            # workbook = xlrd.open_workbook("D:\Jannat USb\Jannat\Republique Francaise")
            # sheet = workbook.sheet_by_index(5)
            # col_val = sheet.col_values(1,1,None) # the second column
            # maxval = max(col_val)
            # print(maxval)
            # maxpos = col_val.index(maxval)+2
            # maxtime = sheet.cell_value(maxpos,0)
            # print(maxtime)

            # # datapath = r'D:\Jannat USb\Jannat\Republique Francaise\Republique_Francaise.xls'
            # # workbook = xlrd.open_workbook(datapath)
            # # # book = xlrd.open_workbook('Republique_Francaise.xlsx')
            # # sheet = workbook.sheet_by_index(0)
            # # col = 5
            # # print(max(sheet.col_values(col)))

            # df = pd.read_excel('Republique_Francaise.xlsx')
            # k=df.loc[df['Domain Name'] == "Achat / Comptabilité / Gestion"]
            # sheet = k.sort_index(5)
            # col_val = sheet.col_values(1,1,None) # the second column
            # maxval = max(col_val)
            # print(maxval)
            # print(k)
            # # print(k.index["Production Date"].max())
            # print(max(k['Production Date']))
            # print (k.index.max())
            # df2 = k.agg(Minimum_Date=('Production Date', np.min), Maximum_Date=('Production Date', np.max))
            # print(df2)
            date_time = datetime.datetime.now()
            date_time=date_time.strftime("%d/%m/%y")
            fileTime=date_time.replace(":","").replace("/","")

            filterbtn=driver.find_elements_by_xpath("//button[@class='btn-clean advanced-search-container__btn ng-tns-c143-3 ng-star-inserted'] | //button[@class='btn-clean advanced-search-container__btn ng-tns-c153-4 ng-star-inserted'] | //*[@id='mainContent']/ng-component/section[1]/app-search-engine/form/div/button[2] | //button[contains(@class,'btn-clean advanced-search-container__btn')]")
            if len(filterbtn)>0:
                filterbtn[0].click()
                time.sleep(2)
                cookiesbtn=driver.find_elements_by_xpath("//aside[@class='gdpr_banner__34yLN']/footer/button[@class='gdpr_banner__subText__JRVbH gdpr_banner__button__1eVD_']")
                if len(cookiesbtn)>0:
                    cookiesbtn[0].click()
                professionalDomainBtn=driver.find_elements_by_xpath("//app-dropdown/mat-form-field/div/div/div/mat-select[contains(@class,'mat-select')]")
                # professionalDomainBtn=driver.find_elements_by_xpath("//div[@class='advanced-search-container__element ng-tns-c143-3 ng-star-inserted']/app-dropdown/mat-form-field/div/div/div/mat-select | //div[@class='advanced-search-container__element ng-tns-c153-4 ng-star-inserted']/app-dropdown/mat-form-field/div/div/div/mat-select | //*[@id='mat-select-4']")
                if len(professionalDomainBtn)>0:
                    professionalDomainBtn[1].click()
                    time.sleep(2)
                    domainDropdown=driver.find_elements_by_xpath("//div[@id='cdk-overlay-0']/div/div/mat-option/mat-pseudo-checkbox | //div[@id='cdk-overlay-1']/div/div/mat-option/mat-pseudo-checkbox")
                    if len(domainDropdown)>0:
                        count=0
                        domainDropdown[count].click()
                        domaindropdownNames=driver.find_elements_by_xpath("//div[@id='cdk-overlay-0']/div/div/mat-option/span[@class='mat-option-text'] | //div[@id='cdk-overlay-1']/div/div/mat-option/mat-pseudo-checkbox")
                        domainName=domaindropdownNames[count].text
                        SearchBtn=driver.find_elements_by_xpath("//button[@class='search-engine__button btn-clean btn ng-tns-c143-3 ng-star-inserted'] | //button[@class='search-engine__button btn-clean btn ng-tns-c153-4 ng-star-inserted'] | //*[@id='mainContent']/ng-component/section[1]/app-search-engine/form/div[1]/button[2]")
                        if len(SearchBtn)>0:
                            SearchBtn[0].submit()
                            driver.refresh()
                        # try:
                        file2 = open("logged "+str(fileTime)+".txt","w")           
                        file2.close()
                        while count < len(domainDropdown):
                            wb_ReadObj=load_workbook("Republique_Francaise.xlsx")
                            sheetObj=wb_ReadObj.active
                            max_Rows=sheetObj.max_row
                            df = pd.read_excel('Republique_Francaise.xlsx')
                            # listofJobsFunction()

                            pLink=driver.find_elements_by_xpath("//ul[@class='pagination ng-star-inserted']/li[@class='pagination__item ng-star-inserted']//a[@class='mat-focus-indicator pagination__item_link mat-button mat-button-base']")
                            if len(pLink)>0:
                                pLink=driver.find_element_by_xpath("//ul[@class='pagination ng-star-inserted']/li[@class='pagination__item ng-star-inserted'][last()]//a[@class='mat-focus-indicator pagination__item_link mat-button mat-button-base'][1]").get_attribute("href")
                                pgnumber=pLink.split('=')
                                pgnumber=pgnumber[-1]
                                pLink=pLink.replace(pgnumber,"")
                                comparenum=1
                            else:
                                comparenum=0
                                pgnumber=0
                            jobCount=0
                            while(comparenum)<=int(pgnumber):
                                if len(pLink)>0:
                                    driver.get(pLink + str(comparenum))

                                # # # listofJobsFunction()
                                AllDomainIndex=df[df["Domain Name"]==domainName].index.values
                                if len(AllDomainIndex)==0:
                                    max_Rows=sheetObj.max_row
                                    TopIndex=max_Rows+1
                                else:
                                    TopIndex=AllDomainIndex[1]
                                    TopIndex=TopIndex+1
                                ListofJobs=driver.find_elements_by_xpath("//ul[@class='list__body ng-tns-c144-2 ng-star-inserted']/li/app-job-list-card/article/div[2] | //ul[@class='list__body ng-tns-c154-3 ng-star-inserted']/li/app-job-list-card/article/div[2] | //ul/li/app-job-list-card/article/div[2]")
                                if len(ListofJobs)>0:
                                    liCount=0
                                    for x in ListofJobs:
                                        jobName=x.find_element_by_xpath(".//div[@class='card-job__content']/div/h3").text
                                        DateofPublication=x.find_element_by_xpath(".//div[@class='card-job__content']/div/p[@class='content__infos__date ng-star-inserted']").text
                                        DateofPublication=DateofPublication.replace("Publié le ","").replace("Posted on ","")
                                        link=x.find_element_by_xpath(".//a[@class='card-job__button btn-clean btn ng-star-inserted']").get_attribute("href")
                                        code=link.split('/')
                                        if len(code)>0:
                                            jobCode=code[-1]
                                        # Location=x.find_element_by_xpath(".//div[@class='card-job__content']/ul/li[1]/app-chip").text.replace("-","")
                                        # l=Location.split(' ')
                                        contractType=x.find_element_by_xpath(".//div[@class='card-job__content']/ul/li[2]/app-chip").text
                                        Location=x.find_element_by_xpath(".//div[@class='card-job__content']/ul/li[1]/app-chip").text.split(' - ')
                                        if len(Location)>0:
                                            if len(Location)>1:
                                                postalCode=Location[0]
                                                city=Location[1]
                                            else:
                                                if(isinstance(Location[0],int)):
                                                    postalCode=Location[0]
                                                    city=""
                                                else:
                                                    city=Location[0]
                                                    postalCode=0
                                        date1=time.strptime(DateofPublication,"%d/%m/%y")
                                        date2=time.strptime(currentDate,"%d/%m/%y")
                                        if(date1)>=date2:
                                            # dfff=df.where[(df["Job Code"]=="116ZVQL") & (df["Domain Name"] == "Achat / Comptabilité / Gestion")].any()
                                            k=df.loc[(df["Job Code"] == jobCode) & (df["Domain Name"] == domainName)].any()
                                            # if(k["Domain Name"]==True):
                                            if(k["Domain Name"]==True):
                                                flag=1
                                                break
                                            else:
                                                # dfff=df.where(df["Domain Name"]==domainName)
                                                # toprow=dfff.iloc[1].index
                                                # dd=df.index.where(df["Domain Name"]==domainName)
                                                # print(dd)

                                                # # # AllDomainIndex=df[df["Domain Name"]==domainName].index.values
                                                # # # TopIndex=AllDomainIndex[0]
                                                # print(ddddd)
                                                # # df.sort_values("Domain Name", inplace=True)
                                                # # topRowCount=df.iloc[1].index
                                                # # print(topRowCount)
                                                # print(df[df["Domain Name"]==domainName].index.values)
                                                # print(dfff)
                                                # sheetObj.insert_rows(int(topRowCount)-1)
                                                # a_row = pd.Series([1, 2])
                                                # row_df = pd.DataFrame([a_row], index = ["row3"])
                                                # df = pd.concat([row_df, df])
                                                # sheetObj.append()


                                                sheetObj.insert_rows(int(TopIndex))
                                                sheetObj.cell(row=TopIndex,column=1).value=domainName
                                                sheetObj.cell(row=TopIndex,column=2).value=jobCode
                                                sheetObj.cell(row=TopIndex,column=3).value=jobName
                                                sheetObj.cell(row=TopIndex,column=5).value=postalCode
                                                sheetObj.cell(row=TopIndex,column=6).value=city
                                                sheetObj.cell(row=TopIndex,column=7).value=DateofPublication
                                                sheetObj.cell(row=TopIndex,column=8).value=link
                                                sheetObj.cell(row=TopIndex,column=9).value=contractType

                                                driver.execute_script("window.open('');")

                                                driver.switch_to.window(driver.window_handles[1])
                                                driver.get(link)

                                                Description=""
                                                CompanyName=""
                                                knowledges=""
                                                Experience=""
                                                Formation=""

                                                Description=driver.find_elements_by_xpath("//p[contains(@class,'job-offer-view__description ng-star-inserted')]")
                                                # Description=driver.find_elements_by_xpath("//p[@class='job-offer-view__description ng-star-inserted'] | //p[@class='app-inner-html--show-list-icons job-offer-view__description ng-star-inserted']")
                                                if len(Description)>0:
                                                    Descriptions=Description[0].text
                                                else:
                                                    Descriptions=''
                                                CompanyName=driver.find_elements_by_xpath("//div[@id='job-offer']/div[@class='ng-star-inserted']/a | //div[@id='job-offer']//a")
                                                if len(CompanyName)>0:
                                                    CompanyName=CompanyName[0].text.replace("\n","").replace("Postuler sur ","").replace("Apply on ","")
                                                else:
                                                    CompanyName=''
                                                infoDivs=driver.find_elements_by_xpath("//div[@class='job-offer-view__details ng-star-inserted']")
                                                if len(infoDivs)>0:
                                                    for x in infoDivs:
                                                        if(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Expérience') | (x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Experience'):
                                                            Experience=x.find_element_by_xpath('./p').text
                                                        elif(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Savoirs et savoir-faire') | (x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Knowledge and know-how'):
                                                            knowledge=x.find_elements_by_xpath('./p')
                                                            if len(knowledge)>0:
                                                                knowledges=knowledge
                                                            knowledge=x.find_elements_by_xpath('./ul/li/p')
                                                            if len(knowledge)>0:
                                                                for j in knowledge: 
                                                                    if len(knowledges)==0:
                                                                        knowledges=j.text
                                                                    else:
                                                                        knowledges=knowledges +','+j.text
                                                            # knowledge=x.find_elements_by_xpath('./p | ./ul/li/p')
                                                            # if len(knowledge)>0:
                                                            #     for j in knowledge:                                                    
                                                            #         knowledges=knowledges +','+j.text
                                                            # else:
                                                            #     knowledges=knowledge

                                                        elif(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Formation') | (x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text=='Training'):
                                                            Formation=x.find_element_by_xpath('./p | ./ul/li/p').text
                                                else:
                                                    Experience=""
                                                    knowledges=""
                                                    Formation=""

                                                sheetObj.cell(row=TopIndex,column=4).value=CompanyName
                                                sheetObj.cell(row=TopIndex,column=10).value=Descriptions
                                                sheetObj.cell(row=TopIndex,column=11).value=Experience
                                                sheetObj.cell(row=TopIndex,column=12).value=knowledges
                                                sheetObj.cell(row=TopIndex,column=13).value=Formation

                                                driver.execute_script("window.close('');")
                                                driver.switch_to.window(driver.window_handles[0])
                                            
                                                TopIndex=TopIndex+1
                                                wb_ReadObj.save("Republique_Francaise.xlsx")
                                                jobCount=jobCount+1
                                                liCount=liCount+1
                                        else:
                                            flag=1
                                            break
                                            
                            
                                # if(k["Domain Name"]==True)or(DateofPublication<=currentDate):
                                if(flag==1):
                                    flag=0
                                    break
                                comparenum=comparenum+1
                            fileforDates = open("logged "+str(fileTime)+".txt","a")           
                            # fileforDates=open("logged14072021.txt","a")
                            fileforDates.writelines("\n Domain Name: "+domainName)
                            fileforDates.writelines("\n Jobs Count: "+str(jobCount))
                            fileforDates.close()
                            jobCount=0


                            filterbtn=driver.find_elements_by_xpath("//button[@class='btn-clean advanced-search-container__btn ng-tns-c143-3 ng-star-inserted'] | //button[@class='btn-clean advanced-search-container__btn ng-tns-c153-4 ng-star-inserted'] | //*[@id='mainContent']/ng-component/section[1]/app-search-engine/form/div/button[2] | //button[contains(@class,'btn-clean advanced-search-container__btn')]")
                            time.sleep(2)
                            if len(filterbtn)>0:
                                filterbtn[0].click()
                                time.sleep(2)
                                cookiesbtn=driver.find_elements_by_xpath("//aside[@class='gdpr_banner__34yLN']/footer/button[@class='gdpr_banner__subText__JRVbH gdpr_banner__button__1eVD_']")
                                if len(cookiesbtn)>0:
                                    cookiesbtn[0].click()
                                professionalDomainBtn=driver.find_elements_by_xpath("//app-dropdown/mat-form-field/div/div/div/mat-select[contains(@class,'mat-select')]")
                                # professionalDomainBtn=driver.find_elements_by_xpath("//div[@class='advanced-search-container__element ng-tns-c143-3 ng-star-inserted']/app-dropdown/mat-form-field/div/div/div/mat-select | //div[@class='advanced-search-container__element ng-tns-c153-4 ng-star-inserted']/app-dropdown/mat-form-field/div/div/div/mat-select | //*[@id='mat-select-4']")
                                if len(professionalDomainBtn)>0:
                                    time.sleep(1)
                                    professionalDomainBtn[1].click()
                                    domainDropdown=driver.find_elements_by_xpath("//div[@id='cdk-overlay-0']/div/div/mat-option/mat-pseudo-checkbox | //div[@id='cdk-overlay-1']/div/div/mat-option/mat-pseudo-checkbox")
                                    if len(domainDropdown)>0:
                                        time.sleep(2)
                                        domainDropdown[count].click()
                                        count=count+1
                                        if count < len(domainDropdown):
                                            time.sleep(1)
                                            domainDropdown[count].click()
                                            domaindropdownNames=driver.find_elements_by_xpath("//div[@id='cdk-overlay-0']/div/div/mat-option/span[@class='mat-option-text'] | //div[@id='cdk-overlay-1']/div/div/mat-option/span[@class='mat-option-text']")
                                            domainName=domaindropdownNames[count].text
                                            SearchBtn=driver.find_elements_by_xpath("//button[@class='search-engine__button btn-clean btn ng-tns-c143-3 ng-star-inserted'] | //button[@class='search-engine__button btn-clean btn ng-tns-c153-4 ng-star-inserted'] | //*[@id='mainContent']/ng-component/section[1]/app-search-engine/form/div[1]/button[2]")
                                            time.sleep(1)
                                            if len(SearchBtn)>0:
                                                SearchBtn[0].submit()
                                                driver.refresh()
                                        else:
                                            break
                        wb_ReadObj.Close()
                        fileforDate = open("DateFile.txt","w+")
                        fileforDate.writelines("Date:"+date_time)
                        fileforDate.close()
                        driver.close()                
    except Exception as e:
        ExceptionFile = open("ExceptionFile.txt","a")
        ExceptionFile.write("Exception in Updating Jobs: "+str(e)+"\n")
        ExceptionFile.close()
        DetailPagesCount = open("DetailPagesCount.txt","w")
        DetailPagesCount.write("Total Detail Pages Complete:"+str(i)+"\n")
        DetailPagesCount.write("Total Rows in Excel:"+str(TotalRow)+"\n")
        DetailPagesCount.close()
        wb_ReadObj.save("Republique_Francaise.xlsx")
        wb_ReadObj.close()
        driver.close()
                #     fileforDate = open("DateFile.txt","w+")
                #     fileforDate.writelines("Date:"+date_time)
                #     fileforDate.close()
                #     driver.close()
                # except:
                #     driver.close()
        ####### Read File Second  #######
    # wb_ReadObj=load_workbook("Republique_Francaise.xlsx")
    # sheetObj=wb_ReadObj.active
    # max_Rows=sheetObj.max_row
    # for i in range(2,max_Rows +1):

    #     cellsData=sheetObj.cell(row=i,column=3)

    #     link=cellsData.value
    #     driver.get(link)
    #     #driver.get('https://www.1jeune1solution.gouv.fr/emplois/116TRRW')
    #     Description=""
    #     CompanyName=""
    #     knowledges=""
    #     Experience=""
    #     Formation=""

    #     Description=driver.find_elements_by_xpath("//p[@class='job-offer-view__description ng-star-inserted']")
    #     if len(Description)>0:
    #         Descriptions=Description[0].text
    #     else:
    #         Descriptions=''
    #     CompanyName=driver.find_elements_by_xpath("//div[@id='job-offer']/div[@class='ng-star-inserted']/a | //div[@id='job-offer']//a")
    #     if len(CompanyName)>0:
    #         CompanyName=CompanyName[0].text.replace("\n","").replace("Postuler sur ","")
    #     else:
    #         CompanyName=''
    #     infoDivs=driver.find_elements_by_xpath("//div[@class='job-offer-view__details ng-star-inserted']")
    #     if len(infoDivs)>0:
    #         for x in infoDivs:
    #             if(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text)=='Expérience':
    #                 Experience=x.find_element_by_xpath('./p').text
    #             elif(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text)=='Savoirs et savoir-faire':
    #                 knowledge=x.find_elements_by_xpath('./p | ./ul/li/p')
    #                 if len(knowledge)>0:
    #                     for j in knowledge:
    #                         knowledges=knowledges +','+j.text
    #                 else:
    #                     print("")

    #             elif(x.find_element_by_xpath(".//h2[@class='job-offer-view__details__title']").text)=='Formation':
    #                 Formation=x.find_element_by_xpath('./p | ./ul/li/p').text
    #     else:
    #         Experience=""
    #         knowledges=""
    #         Formation=""

    #     sheetObj.cell(row=i,column=8).value=Descriptions
    #     sheetObj.cell(row=i,column=9).value=CompanyName
    #     sheetObj.cell(row=i,column=10).value=Experience
    #     sheetObj.cell(row=i,column=11).value=knowledges
    #     sheetObj.cell(row=i,column=12).value=Formation


    #     wb_ReadObj.save("Republique_Francaise.xlsx")
    #     print(CompanyName,"     ",)
    #     time.sleep(3)




    ###### Update Code #########







